import os
import json
from openpyxl import Workbook, load_workbook

def handler(event, context):
    token = event['queryStringParameters'].get('token')
    data_dir = 'data'
    os.makedirs(data_dir, exist_ok=True)
    
    file_path = f'{data_dir}/{token}.xlsx'
    
    if event['httpMethod'] == 'GET':
        if not os.path.exists(file_path):
            return {'statusCode': 404}
            
        wb = load_workbook(file_path)
        ws = wb.active
        links = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            links.append({
                'url': row[0],
                'category': row[1],
                'hashtags': row[2].split(','),
                'pinned': row[3]
            })
        return {'statusCode': 200, 'body': json.dumps(links)}
    
    elif event['httpMethod'] == 'POST':
        wb = Workbook() if not os.path.exists(file_path) else load_workbook(file_path)
        ws = wb.active
        if ws.max_row == 1:
            ws.append(['URL', 'Category', 'Hashtags', 'Pinned'])
        
        link_data = json.loads(event['body'])
        ws.append([
            link_data['url'],
            link_data['category'],
            ','.join(link_data['hashtags']),
            link_data['pinned']
        ])
        
        wb.save(file_path)
        return {'statusCode': 200}
